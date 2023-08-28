# coding=gb18030


import xlrd
import numpy
import openpyxl
import time
from gurobipy import *

MAX_LABEL_COST = 1000000


g_node_list = []
g_link_list = []
g_bus_list = []
g_state_vector_list = []

g_number_of_nodes = 0
g_number_of_links = 0
g_number_of_buses = 0
g_number_of_state_vectors = 0

# need to pre determined
g_number_of_destinations = 3 + 1
g_number_of_time_intervals = 60 + 1

     
class Node:
    def __init__(self):
        self.node_id = 0
        self.ingoing_node_list = []
        self.ingoing_link_list = []
        self.outgoing_node_list = []
        self.outgoing_link_list = []


class Link:
    def __init__(self):
        self.link_id = 0
        self.from_node_id = 0
        self.to_node_id = 0
        self.type = 0
        self.cost = 0
        self.travel_time = 0
        self.demand_vector_id = 0
        self.start_time = 0
        self.end_time = 0
        self.LR_multiplier = []
        self.outgoing_from_depot_flag = 0

class Bus:
    def __init__(self):
        self.bus_id = 0
        self.from_node_id = 0
        self.to_node_id = 0
        self.departure_time = 0
        self.arrival_time = 0
        self.initial_state_vector_id = 0
        self.final_state_vector_id = 0
        self.cap = 0
        self.node_sequence = []
        self.time_sequence = []
        self.link_sequence = []
        self.state_vector_sequence = []
        self.available_node_list = []
        self.available_link_list = []
        self.available_state_vector_list = []
        self.time_state_dependent_link_volume = []
        self.time_state_dependent_link_from_state = []
        self.time_state_dependent_link_to_state = []
        self.time_state_dependent_link_cost = []
        self.reemployed_bus_flag = 0
        

        
class State_Vector:
     def __init__(self):
         self.state_vector_id = 0
         self.passenger_by_destination = []
         self.total_number_of_passengers = 0
         self.to_state_list_on_pickup_arc = []
         self.state_dependent_cost = 0
         self.to_state_list_on_transfer_arc = []         
  
        

def g_read_input_data():    
    
    #initialization

    global g_number_of_nodes
    global g_number_of_links
    global g_number_of_buses
    global g_number_of_state_vectors


    # read node
    file = xlrd.open_workbook("input_node.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    node = Node()
    node.node_id = 0
    g_node_list.append(node)
    g_number_of_nodes += 1    
    for row in range(1, sheet.nrows):
        try:
            node = Node()
            node.node_id = int(sheet.cell_value(row,0))
            g_node_list.append(node)
            g_number_of_nodes += 1
            print('reading {} passenger nodes..'.format(g_number_of_nodes))
        except:
            print('Read error. Check your passenger node file')
    print('passenger nodes_number:{}'.format(g_number_of_nodes))  
    
    # read links
    file = xlrd.open_workbook("input_link.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    link = Link()
    link.link_id = 0
    g_link_list.append(link)
    g_number_of_links += 1
    for row in range(1, sheet.nrows):
        try:
            link = Link()
            link.link_id = int(sheet.cell_value(row, 0))
            link.from_node_id = int(sheet.cell_value(row, 1))
            link.to_node_id = int(sheet.cell_value(row, 2))
            link.type = int(sheet.cell_value(row, 3))
            link.cost = float(sheet.cell_value(row, 4))
            link.travel_time = int(sheet.cell_value(row, 5))
            link.demand_vector_id = int(sheet.cell_value(row, 6))
            link.start_time = int(sheet.cell_value(row, 7))
            link.end_time = int(sheet.cell_value(row, 8))
            link.outgoing_from_depot_flag = int(sheet.cell_value(row, 9))
            g_link_list.append(link)
            g_number_of_links += 1
            print('reading {} passenger links..'.format(g_number_of_links))
        except:
            print('Read error. Check your passenger link file')
    print('passenger links_number:{}'.format(g_number_of_links))
    
    # read buses
    file = xlrd.open_workbook("input_bus.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    bus = Bus()
    bus.bus_id = 0
    g_bus_list.append(bus)
    g_number_of_buses += 1
    for row in range(1, sheet.nrows):
        try:
            bus = Bus()
            bus.bus_id = int(sheet.cell_value(row, 0))
            bus.from_node_id = int(sheet.cell_value(row, 1))
            bus.to_node_id = int(sheet.cell_value(row, 2))
            bus.departure_time = int(sheet.cell_value(row, 3))
            bus.arrival_time = int(sheet.cell_value(row, 4))
            bus.initial_state_vector_id = int(sheet.cell_value(row, 5))  
            bus.final_state_vector_id = int(sheet.cell_value(row, 6))            
            bus.cap = int(sheet.cell_value(row, 7))
            available_node_list = str(sheet.cell_value(row, 8))
            bus.available_node_list = available_node_list.strip().split(';')
            # transfer str to int
            bus.available_node_list = [int (node) for node in bus.available_node_list]
            available_link_list = str(sheet.cell_value(row, 9))
            bus.available_link_list = available_link_list.strip().split(';')
            # transfer str to int
            bus.available_link_list = [int (link) for link in bus.available_link_list]
            available_state_vector_list = str(sheet.cell_value(row, 11))
            bus.available_state_vector_list = available_state_vector_list.strip().split(';')
            # transfer str to int
            bus.available_state_vector_list = [int (state) for state in bus.available_state_vector_list]            
            bus.reemployed_bus_flag = int(sheet.cell_value(row, 10))
            g_bus_list.append(bus)
            g_number_of_buses += 1
            print('reading {} buses..'.format(g_number_of_buses))
        except:
            print('Read. Check your passenger file')
    print('buses_number:{}'.format(g_number_of_buses))  
    
    

    # read state column
    file = xlrd.open_workbook("input_state_vector.xlsx") # open file
    sheet = file.sheet_by_index(0) # open sheet
    state = State_Vector()
    state.state_id = 0
    g_state_vector_list.append(state)
    g_number_of_state_vectors += 1
    for row in range(1, sheet.nrows):
        try:
            state = State_Vector()
            state.state_vector_id = int(sheet.cell_value(row, 0))
            passenger_by_destination = str(sheet.cell_value(row, 1))
            state.passenger_by_destination = passenger_by_destination.strip().split(';')
            # transfer str to int
            state.passenger_by_destination = [int (passenger) for passenger in state.passenger_by_destination]
            state.total_number_of_passengers = int(sheet.cell_value(row, 2))
            to_state_list_on_pickup_arc = str(sheet.cell_value(row, 3))
            state.to_state_list_on_pickup_arc = to_state_list_on_pickup_arc.strip().split(';')
            # transfer str to int
            state.to_state_list_on_pickup_arc = [int (state) for state in state.to_state_list_on_pickup_arc]
            state.state_dependent_cost = int(sheet.cell_value(row, 4))
            to_state_list_on_transfer_arc = str(sheet.cell_value(row, 5))
            state.to_state_list_on_transfer_arc = to_state_list_on_transfer_arc.strip().split(';')
            # transfer str to int
            state.to_state_list_on_transfer_arc = [int (state) for state in state.to_state_list_on_transfer_arc]
            #state.to_state_list_on_transfer_arc = list(state.to_state_list_on_transfer_arc)
            g_state_vector_list.append(state)
            g_number_of_state_vectors += 1
            print('reading {} state columns..'.format(g_number_of_state_vectors))
        except:
            print('Read. Check your state file')
    print('state_column_number:{}'.format(g_number_of_state_vectors))

            
                         
def  g_generate_in_out_going_link():   

    # record ingoing and outgoing nodes and links for each passenger node       
    for l in range(1, g_number_of_links):
        link_id = g_link_list[l].link_id
        from_node_id = g_link_list[l].from_node_id
        to_node_id = g_link_list[l].to_node_id     
        g_node_list[to_node_id].ingoing_link_list.append(link_id)
        g_node_list[to_node_id].ingoing_node_list.append(from_node_id)
        g_node_list[from_node_id].outgoing_link_list.append(link_id)
        g_node_list[from_node_id].outgoing_node_list.append(to_node_id)
        
    return()


def g_use_Gurobi_to_sovle_the_model():
    
    global link_set
    global time_set
    global state_vector_set
    global bus_set
    global m
    global x
    global y
    
    m = Model("space_time_state_model")
    
    # establish sets
    link_set = []
    time_set = []
    state_vector_set = []
    bus_set = []
    
    
    for l in range(1, g_number_of_links):
        link_id = g_link_list[l].link_id
        link_set.append(link_id)
        
    for t in range(1, g_number_of_time_intervals):
        time_set.append(t)
    
    for s in range(1, g_number_of_state_vectors):
        state_vector_id = g_state_vector_list[s].state_vector_id
        state_vector_set.append(state_vector_id)
        
    for b in range(1, g_number_of_buses):
        bus_id = g_bus_list[b].bus_id
        bus_set.append(bus_id)
    
    # decision variables
    x = m.addVars(bus_set, link_set, time_set, time_set, state_vector_set, state_vector_set, vtype = GRB.BINARY, name = 'x')
    
    # obejctive function
    objective_function = sum(sum(sum(sum(sum(sum(x[bus_id, link_id, from_t, from_t + g_link_list[link_id].travel_time, from_s, to_s] * (g_link_list[link_id].cost + g_state_vector_list[from_s].state_dependent_cost) \
                                             for to_s in state_vector_set) for from_s in state_vector_set) for to_t in time_set if (to_t == from_t + g_link_list[link_id].travel_time)) for from_t in time_set) for link_id in link_set) for bus_id in bus_set)
    
    m.setObjective(objective_function, GRB.MINIMIZE)
    
    # vehicle flow balance consraint
    for b in bus_set:
        bus_id = g_bus_list[b].bus_id
        from_node_id = g_bus_list[bus_id].from_node_id            
        from_t = g_bus_list[bus_id].departure_time
        from_s = 1        
        m.addConstr(sum(sum(sum(x[bus_id, outgoing_link_id, from_t, to_t, from_s, to_s] for to_s in state_vector_set) for to_t in time_set if (to_t == from_t + g_link_list[outgoing_link_id].travel_time)) for outgoing_link_id in g_node_list[from_node_id].outgoing_link_list) == 1, "c1.1")

        for from_s_other in state_vector_set:
            if (from_s_other != from_s):
                m.addConstr(sum(sum(sum(x[bus_id, outgoing_link_id, from_t, to_t, from_s_other, to_s] for to_s in state_vector_set) for to_t in time_set if (to_t == from_t + g_link_list[outgoing_link_id].travel_time)) for outgoing_link_id in g_node_list[from_node_id].outgoing_link_list) == 0, "c1.2")
        
        for n in range(1, g_number_of_nodes):
            node_id_other = g_node_list[n].node_id
            if (node_id_other != from_node_id):
                for from_s_other in state_vector_set:
                    m.addConstr(sum(sum(sum(x[bus_id, outgoing_link_id, from_t, to_t, from_s_other, to_s] for to_s in state_vector_set) for to_t in time_set if (to_t == from_t + g_link_list[outgoing_link_id].travel_time)) for outgoing_link_id in g_node_list[node_id_other].outgoing_link_list) == 0, "c1.3")
                    
        
    for b in bus_set:
        bus_id = g_bus_list[b].bus_id
        to_node_id = g_bus_list[bus_id].to_node_id            
        to_t = g_bus_list[bus_id].arrival_time
        to_s = 1
        m.addConstr(sum(sum(sum(x[bus_id, ingoing_link_id, from_t, to_t, from_s, to_s] for from_s in state_vector_set) for from_t in time_set if (from_t == to_t -  g_link_list[ingoing_link_id].travel_time)) for ingoing_link_id in g_node_list[to_node_id].ingoing_link_list) == 1, "c2")

    for b in bus_set:
        bus_id = g_bus_list[b].bus_id
        for n in range(1, g_number_of_nodes):
            node_id = g_node_list[n].node_id
            for key_t in time_set:  
                if (key_t >= 2) and (key_t <= 9):
                    for key_s in state_vector_set:
                        m.addConstr(sum(sum(sum(x[bus_id, outgoing_link_id, key_t, to_t, key_s, to_s] for to_s in state_vector_set) for to_t in time_set if (to_t == key_t + g_link_list[outgoing_link_id].travel_time) ) for outgoing_link_id in g_node_list[node_id].outgoing_link_list) == \
                                    sum(sum(sum(x[bus_id, ingoing_link_id, from_t, key_t, from_s, key_s] for from_s in state_vector_set) for from_t in time_set if (from_t == key_t - g_link_list[ingoing_link_id].travel_time) ) for ingoing_link_id in g_node_list[node_id].ingoing_link_list), "c3")
    
    # demand satisfiction constraint
    for l in range(1, g_number_of_links):
        link_id = g_link_list[l].link_id
        if (g_link_list[link_id].type == 1):
            demand_vector_id = g_link_list[l].demand_vector_id
            start_time =  g_link_list[l].start_time
            end_time =  g_link_list[l].end_time
            for d in range(1, g_number_of_destinations):
                m.addConstr(sum(sum(sum(sum(sum(x[bus_id, link_id, from_t, to_t, from_s, to_s] * (g_state_vector_list[to_s].passenger_by_destination[d] - g_state_vector_list[from_s].passenger_by_destination[d]) for to_s in state_vector_set) for from_s in state_vector_set) \
                                        for to_t in time_set if (to_t == from_t + g_link_list[link_id].travel_time) ) for from_t in time_set if (from_t >= start_time and from_t <= end_time)) for bus_id in bus_set) == g_state_vector_list[demand_vector_id].passenger_by_destination[d], "c4")
    
  
    # transfer constraint
    for l in range(1, g_number_of_links):
        link_id = g_link_list[l].link_id
        if (g_link_list[link_id].type == 2):
            for current_bus_id in bus_set:
                for from_t in time_set:
                    to_t = from_t + g_link_list[link_id].travel_time
                    if (to_t in time_set):
                        for d in range(1, g_number_of_destinations):
                            change_of_passenger_by_current_bus = sum(sum(x[current_bus_id, link_id, from_t, to_t, from_s, to_s] * (g_state_vector_list[from_s].passenger_by_destination[d] - g_state_vector_list[to_s].passenger_by_destination[d]) for to_s in state_vector_set) for from_s in state_vector_set)
                            change_of_passenger_by_other_bus = sum(sum(sum(x[bus_id_other, link_id, from_t, to_t, from_s, to_s] * (g_state_vector_list[to_s].passenger_by_destination[d] - g_state_vector_list[from_s].passenger_by_destination[d]) for to_s in state_vector_set) for from_s in state_vector_set) for bus_id_other in bus_set if (bus_id_other != current_bus_id))
                            m.addConstr(change_of_passenger_by_current_bus == change_of_passenger_by_other_bus, "c5")



    # state transition constraint
    for l in range(1, g_number_of_links):
        link_id = g_link_list[l].link_id 
        if (g_link_list[link_id].type == 1):
            for bus_id in bus_set:
                for from_t in time_set: 
                    for to_t in time_set:
                        if (to_t == from_t + g_link_list[link_id].travel_time):
                            for from_s in state_vector_set:
                                for to_s in state_vector_set:
                                    for d in range(1, g_number_of_destinations):
                                        if (g_state_vector_list[to_s].passenger_by_destination[d] < g_state_vector_list[from_s].passenger_by_destination[d]):
                                            m.addConstr(x[bus_id, link_id, from_t, to_t, from_s, to_s] == 0, "c6")
        if (g_link_list[link_id].type == 0):
            for bus_id in bus_set:
                for from_t in time_set:
                    for to_t in time_set:
                        if (to_t == from_t + g_link_list[link_id].travel_time):
                            for from_s in state_vector_set:
                                for to_s in state_vector_set:
                                    if (to_s != from_s):
                                        m.addConstr(x[bus_id, link_id, from_t, to_t, from_s, to_s] == 0, "c7")
        if (g_link_list[link_id].type == -1):
            for bus_id in bus_set:
                for from_t in time_set:
                    for to_t in time_set:
                        if (to_t == from_t + g_link_list[link_id].travel_time):
                            for from_s in state_vector_set:
                                for to_s in state_vector_set:
                                    for d in range(1, g_number_of_destinations):
                                        if (d == 1):
                                            if (g_state_vector_list[to_s].passenger_by_destination[d] != 0):
                                                m.addConstr(x[bus_id, link_id, from_t, to_t, from_s, to_s] == 0, "c8")

                
                                
                                
    m.setParam('MIPGap', 0)
    m.optimize()    
    if m.status == GRB.OPTIMAL:
        m.printAttr('X')
    if m.status == GRB.Status.INFEASIBLE:
        print('Optimization was stopped with status %d' % m.status)
        # do IIS, find infeasible constraints
        m.computeIIS()
        for c in m.getConstrs():
            if c.IISConstr:
                print('%s' % c.constrName)
                    
    m.write("m.lp")
    
    for b in bus_set:
        bus_id = g_bus_list[b].bus_id
        for l in range(1, g_number_of_links):
            link_id = g_link_list[l].link_id
            for t in time_set:
                if (t + g_link_list[link_id].travel_time <= g_number_of_time_intervals - 1):
                    for from_s in state_vector_set:
                        for to_s in state_vector_set:
                            current_variable_value = x[bus_id, link_id, t, t + g_link_list[link_id].travel_time, from_s, to_s].x
                            if (current_variable_value == 1):
                                print(x[bus_id, link_id, t, t + g_link_list[link_id].travel_time, from_s, to_s])
    

    
    
    return()                 
                 
                    
def g_time_dependent_dynamic_programming_for_buses():
    
    global to_s_destination_number
    global from_s_origin_number

    # dynamic programming for each passenger
    for re_assignment_iteration in range(1, 5):
        for p in range(1, g_number_of_buses):
            # initial
            g_bus_list[p].time_state_dependent_link_volume = [[[0 for s in range(0, g_number_of_state_vectors + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
            g_bus_list[p].time_state_dependent_link_from_state = [[[0 for s in range(0, g_number_of_state_vectors + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
            g_bus_list[p].time_state_dependent_link_to_state = [[[0 for s in range(0, g_number_of_state_vectors + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
            g_bus_list[p].time_state_dependent_link_cost = [[[[100000 for s2 in range(0, g_number_of_state_vectors + 1)] for s1 in range(0, g_number_of_state_vectors + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
            
            # update demand satisfaction
            g_update_satisfied_demand_by_other_buses()  
            g_update_transfer_demand_generated_by_other_buses()
            
            pre_node_id = - numpy.ones([g_number_of_nodes, g_number_of_time_intervals, g_number_of_state_vectors])
            pre_time_interval = - numpy.ones([g_number_of_nodes, g_number_of_time_intervals, g_number_of_state_vectors])
            pre_state_vector_id = - numpy.ones([g_number_of_nodes, g_number_of_time_intervals, g_number_of_state_vectors])
            label_cost = MAX_LABEL_COST * numpy.ones([g_number_of_nodes, g_number_of_time_intervals, g_number_of_state_vectors])
            
            g_bus_list[p].node_sequence = []
            g_bus_list[p].time_sequence = []   
            g_bus_list[p].state_vector_sequence = []
            g_bus_list[p].link_sequence = []
        
            # get agent information
            from_node_id = int (g_bus_list[p].from_node_id)
            to_node_id = int (g_bus_list[p].to_node_id)
            departure_time = int (g_bus_list[p].departure_time)
            arrival_time = int (g_bus_list[p].arrival_time)
            initial_state_vector_id = int (g_bus_list[p].initial_state_vector_id)
            final_state_vector_id = int (g_bus_list[p].final_state_vector_id)
                                
            # initialize
            pre_node_id[from_node_id][departure_time][initial_state_vector_id] = 0
            pre_time_interval[from_node_id][departure_time][initial_state_vector_id] = departure_time
            pre_state_vector_id[from_node_id][departure_time][initial_state_vector_id] = 0
            label_cost[from_node_id][departure_time][initial_state_vector_id] = 0
            
            reemployed_bus_flag = g_bus_list[p].reemployed_bus_flag
            
            # dynamic programming
            for t in range(departure_time, arrival_time):        
                for n in g_bus_list[p].available_node_list:                           
                    # if (pre_node_id[n][t] != -1):
                    for l in range(0, len(g_node_list[n].outgoing_link_list)):
                        outgoing_link_id = g_node_list[n].outgoing_link_list[l]
                        outgoing_node_id = g_node_list[n].outgoing_node_list[l]
                        outgoing_from_depot_flag = g_link_list[outgoing_link_id].outgoing_from_depot_flag
                        if (outgoing_link_id in g_bus_list[p].available_link_list):
                            fixed_cost = g_link_list[outgoing_link_id].cost
                            if (outgoing_from_depot_flag == 1):
                                if (reemployed_bus_flag == 1):
                                    fixed_cost += usage_cost_of_reemployed_shuttles
                            demand_vector_id = g_link_list[outgoing_link_id].demand_vector_id
                            trave_time = g_link_list[outgoing_link_id].travel_time
                            if (g_link_list[outgoing_link_id].type == 1):
                                start_time = g_link_list[outgoing_link_id].start_time
                                end_time = g_link_list[outgoing_link_id].end_time
                                if (start_time <= t <= end_time):
                                    for from_s in range(1, g_number_of_state_vectors):
                                        state_dependent_cost = g_state_vector_list[from_s].state_dependent_cost
                                        if (pre_node_id[n][t][from_s] != -1):
                                            for to_s in range(1, g_number_of_state_vectors):
                                                if (to_s in g_state_vector_list[from_s].to_state_list_on_pickup_arc):
                                                    deviation_state_vector = [g_state_vector_list[to_s].passenger_by_destination[d] - g_state_vector_list[from_s].passenger_by_destination[d] for d in range(0, g_number_of_destinations)]
                                                    demand_vector = g_state_vector_list[demand_vector_id].passenger_by_destination
                                                    if (deviation_state_vector <= demand_vector):
                                                        LR_multiplier = g_link_list[outgoing_link_id].LR_multiplier[iteration_step]
                                                        ADMM_multiplier = 0
                                                        total_change_of_state = 0
                                                        # for loop for each destination
                                                        for d in range(1, g_number_of_destinations):
                                                            change_of_state = g_state_vector_list[to_s].passenger_by_destination[d] - g_state_vector_list[from_s].passenger_by_destination[d]
                                                            ADMM_multiplier += 0.5 * rou1 * change_of_state * change_of_state + rou1 * change_of_state * (link_satisfied_demand_by_other_buses[outgoing_link_id][d] - g_state_vector_list[demand_vector_id].passenger_by_destination[d])
                                                            total_change_of_state += change_of_state
                                                        cost = fixed_cost + state_dependent_cost+LR_multiplier * total_change_of_state + ADMM_multiplier
                                                        g_bus_list[p].time_state_dependent_link_cost[outgoing_link_id][t][from_s][to_s] = cost
                                                        if (t + trave_time < g_number_of_time_intervals):
                                                            if (label_cost[n][t][from_s] + cost <= label_cost[outgoing_node_id][t + trave_time][to_s]):
                                                                label_cost[outgoing_node_id][t + trave_time][to_s] = label_cost[n][t][from_s] + cost
                                                                pre_node_id[outgoing_node_id][t + trave_time][to_s] = n
                                                                pre_time_interval[outgoing_node_id][t + trave_time][to_s] = t
                                                                pre_state_vector_id[outgoing_node_id][t + trave_time][to_s] = from_s
                                                                
                                else:
                                    for from_s in range(1, g_number_of_state_vectors):
                                        state_dependent_cost = g_state_vector_list[from_s].state_dependent_cost
                                        if (pre_node_id[n][t][from_s] != -1):
                                            to_s = from_s 
                                            cost = fixed_cost + state_dependent_cost
                                            g_bus_list[p].time_state_dependent_link_cost[outgoing_link_id][t][from_s][to_s] = cost
                                            if (t + trave_time < g_number_of_time_intervals):
                                                if (label_cost[n][t][from_s] + cost <= label_cost[outgoing_node_id][t + trave_time][to_s]):
                                                    label_cost[outgoing_node_id][t + trave_time][to_s] = label_cost[n][t][from_s] + cost
                                                    pre_node_id[outgoing_node_id][t + trave_time][to_s] = n
                                                    pre_time_interval[outgoing_node_id][t + trave_time][to_s] = t
                                                    pre_state_vector_id[outgoing_node_id][t + trave_time][to_s] = from_s 
                                    
                            if (g_link_list[outgoing_link_id].type == 0):
                                for from_s in range(1, g_number_of_state_vectors):
                                    state_dependent_cost = g_state_vector_list[from_s].state_dependent_cost
                                    if (pre_node_id[n][t][from_s] != -1):
                                        to_s = from_s 
                                        cost = fixed_cost + state_dependent_cost
                                        g_bus_list[p].time_state_dependent_link_cost[outgoing_link_id][t][from_s][to_s] = cost
                                        if (t + trave_time < g_number_of_time_intervals):
                                            if (label_cost[n][t][from_s] + cost <= label_cost[outgoing_node_id][t + trave_time][to_s]):
                                                label_cost[outgoing_node_id][t + trave_time][to_s] = label_cost[n][t][from_s] + cost
                                                pre_node_id[outgoing_node_id][t + trave_time][to_s] = n
                                                pre_time_interval[outgoing_node_id][t + trave_time][to_s] = t
                                                pre_state_vector_id[outgoing_node_id][t + trave_time][to_s] = from_s    
                            
                            if (g_link_list[outgoing_link_id].type < 0):
                                destination_id = abs(g_link_list[outgoing_link_id].type)
                                for from_s in range(1, g_number_of_state_vectors):
                                    state_dependent_cost = g_state_vector_list[from_s].state_dependent_cost
                                    if (pre_node_id[n][t][from_s] != -1):
                                        current_state_vector = []
                                        for d in range(0, g_number_of_destinations):
                                            if (d == destination_id):
                                                current_state_vector.append(0)
                                            if (d != destination_id):
                                                current_state_vector.append(g_state_vector_list[from_s].passenger_by_destination[d])    
                                        current_state_vector_element = list(filter(lambda x: x.passenger_by_destination == current_state_vector, g_state_vector_list))[0]
                                        to_s = current_state_vector_element.state_vector_id  
                                        cost = fixed_cost + state_dependent_cost
                                        g_bus_list[p].time_state_dependent_link_cost[outgoing_link_id][t][from_s][to_s] = cost
                                        if (t + trave_time < g_number_of_time_intervals):
                                            if (label_cost[n][t][from_s] + cost <= label_cost[outgoing_node_id][t + trave_time][to_s]):
                                                label_cost[outgoing_node_id][t + trave_time][to_s] = label_cost[n][t][from_s] + cost
                                                pre_node_id[outgoing_node_id][t + trave_time][to_s] = n
                                                pre_time_interval[outgoing_node_id][t + trave_time][to_s] = t
                                                pre_state_vector_id[outgoing_node_id][t + trave_time][to_s] = from_s 
                                                
                            if (g_link_list[outgoing_link_id].type == 2):
#                                for from_s in range(1, g_number_of_state_vectors):
#                                    state_dependent_cost = g_state_vector_list[from_s].state_dependent_cost
#                                    if (pre_node_id[n][t][from_s] != -1):
#                                        for to_s in g_state_vector_list[from_s].to_state_list_on_transfer_arc:                             
#                                            LR_multiplier = g_link_list[outgoing_link_id].LR_multiplier[iteration_step]
#                                            #LR_multiplier = 0
#                                            ADMM_multiplier = 0
#                                            total_change_of_state = 0
#                                            #from_passegner_numnber = g_state_vector_list[from_s].total_number_of_passengers
#                                            #to_passegner_numnber = g_state_vector_list[to_s].total_number_of_passengers
#                                            #if  g_state_vector_list[from_s].state_vector_id != g_state_vector_list[to_s].state_vector_id:
#                                            total_transfer_penalty = 1*transfer_penalty
#                                            # for loop for each destination
#                                            for d in range(1, g_number_of_destinations):
#                                                change_of_state = g_state_vector_list[to_s].passenger_by_destination[d] - g_state_vector_list[from_s].passenger_by_destination[d] 
#                                                ADMM_multiplier += 0.5 * rou2 * change_of_state * change_of_state - rou2 * change_of_state * link_transfer_demand_generated_by_other_buses[outgoing_link_id][t][d]
#                                                total_change_of_state += change_of_state
#                                            cost = fixed_cost  + state_dependent_cost  + total_transfer_penalty + LR_multiplier * total_change_of_state + ADMM_multiplier
#                                            g_bus_list[p].time_state_dependent_link_cost[outgoing_link_id][t][from_s][to_s] = cost
#                                            if (t + trave_time < g_number_of_time_intervals):
#                                                if (label_cost[n][t][from_s] + cost <= label_cost[outgoing_node_id][t + trave_time][to_s]):
#                                                    label_cost[outgoing_node_id][t + trave_time][to_s] = label_cost[n][t][from_s] + cost
#                                                    pre_node_id[outgoing_node_id][t + trave_time][to_s] = n
#                                                    pre_time_interval[outgoing_node_id][t + trave_time][to_s] = t
#                                                    pre_state_vector_id[outgoing_node_id][t + trave_time][to_s] = from_s
                                to_s_destination_number = [0,0,0]
                                from_s_origin_number = [0,0,0]
                                start_time = g_link_list[outgoing_link_id].start_time
                                end_time = g_link_list[outgoing_link_id].end_time
                                if (start_time <= t <= end_time):
                                    for from_s in range(1, g_number_of_state_vectors):
                                        state_dependent_cost = g_state_vector_list[from_s].state_dependent_cost
                                        if (pre_node_id[n][t][from_s] != -1):
                                            for to_s in range(1, g_number_of_state_vectors):
                                                if (to_s in g_state_vector_list[from_s].to_state_list_on_transfer_arc):
                                                    transfer_state_vector = g_state_vector_list[to_s].state_dependent_cost
                                                    LR_multiplier = g_link_list[outgoing_link_id].LR_multiplier[iteration_step]
#                                                    LR_multiplier = 0
                                                    ADMM_multiplier = 0
                                                    total_change_of_state = 0
                                                    # for loop for each destination
#                                                    for d in range(1, g_number_of_destinations):
#                                                        to_s_destination_number[d-1]= to_s_destination_number[d-1] + g_state_vector_list[to_s].passenger_by_destination[d]
#                                                        from_s_origin_number[d-1] = from_s_origin_number[d-1] + g_state_vector_list[from_s].passenger_by_destination[d]                                                       
#                                                    print(to_s_destination_number,from_s_origin_number)
                                                    for d in range(1, g_number_of_destinations):                                                        
#                                                        if to_s_destination_number[d-1] == from_s_origin_number[d-1]:
                                                        change_of_state = g_state_vector_list[to_s].passenger_by_destination[d] - g_state_vector_list[from_s].passenger_by_destination[d]
                                                        ADMM_multiplier += 0.5 * rou2 * change_of_state * change_of_state - rou2 * change_of_state * link_transfer_demand_generated_by_other_buses[outgoing_link_id][t][d]
                                                        total_change_of_state += change_of_state
                                                    cost = fixed_cost + transfer_state_vector + LR_multiplier * total_change_of_state + ADMM_multiplier
                                                    g_bus_list[p].time_state_dependent_link_cost[outgoing_link_id][t][from_s][to_s] = cost
                                                    if (t + trave_time < g_number_of_time_intervals):
                                                        if (label_cost[n][t][from_s] + cost <= label_cost[outgoing_node_id][t + trave_time][to_s]):
                                                            label_cost[outgoing_node_id][t + trave_time][to_s] = label_cost[n][t][from_s] + cost
                                                            pre_node_id[outgoing_node_id][t + trave_time][to_s] = n
                                                            pre_time_interval[outgoing_node_id][t + trave_time][to_s] = t
                                                            pre_state_vector_id[outgoing_node_id][t + trave_time][to_s] = from_s
                                                                
                                else:
                                    for from_s in range(1, g_number_of_state_vectors):
                                        state_dependent_cost = g_state_vector_list[from_s].state_dependent_cost
                                        if (pre_node_id[n][t][from_s] != -1):
                                            to_s = from_s 
                                            cost = fixed_cost + state_dependent_cost
                                            g_bus_list[p].time_state_dependent_link_cost[outgoing_link_id][t][from_s][to_s] = cost
                                            if (t + trave_time < g_number_of_time_intervals):
                                                if (label_cost[n][t][from_s] + cost <= label_cost[outgoing_node_id][t + trave_time][to_s]):
                                                    label_cost[outgoing_node_id][t + trave_time][to_s] = label_cost[n][t][from_s] + cost
                                                    pre_node_id[outgoing_node_id][t + trave_time][to_s] = n
                                                    pre_time_interval[outgoing_node_id][t + trave_time][to_s] = t
                                                    pre_state_vector_id[outgoing_node_id][t + trave_time][to_s] = from_s 



                                           
            # backtrace                    
            n = to_node_id                 
            t = arrival_time
            s = final_state_vector_id
            g_bus_list[p].node_sequence.insert(0, n)
            g_bus_list[p].time_sequence.insert(0, t) 
            g_bus_list[p].state_vector_sequence.insert(0, s)
            
            if (label_cost[n][t][s] == MAX_LABEL_COST):
                print('can not find space-time-state path for bus:{}'.format(p))  
                
            for backtrace_step in range(1, g_number_of_time_intervals):
                if (label_cost[n][t][s] != MAX_LABEL_COST):
                    if (n != from_node_id) or (t != departure_time):
                        pre_n = int (pre_node_id[n][t][s])
                        pre_t = int (pre_time_interval[n][t][s])
                        pre_s = int (pre_state_vector_id[n][t][s])
                        # get current link id
                        link_element = list(filter(lambda x: x.from_node_id == pre_n and x.to_node_id == n, g_link_list))[0]
                        l = link_element.link_id                
                        # l = int (node_link_map.loc[(node_link_map.from_node_id == pre_n) & (node_link_map.to_node_id == n),'link_id'].values[0])
                        # update time-dependent link volume
                        g_bus_list[p].time_state_dependent_link_volume[l][pre_t][pre_s] += 1
                        g_bus_list[p].time_state_dependent_link_from_state[l][pre_t][pre_s] = pre_s
                        g_bus_list[p].time_state_dependent_link_to_state[l][pre_t][pre_s] = s
                        n = pre_n
                        t = pre_t 
                        s = pre_s
                        # node and time sequence lists
                        g_bus_list[p].node_sequence.insert(0, n)
                        g_bus_list[p].time_sequence.insert(0, t)
                        g_bus_list[p].state_vector_sequence.insert(0, s)
                        g_bus_list[p].link_sequence.insert(0, l)
         

    return()



def g_initialization():
    
    for l in range(1, g_number_of_links):
        g_link_list[l].LR_multiplier = [0.01 for l in range(1, maximum_iteration_step + 2)]
    
    for p in range(1, g_number_of_buses):
        # initial
        g_bus_list[p].time_state_dependent_link_volume = [[[0 for s in range(0, g_number_of_state_vectors + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
        g_bus_list[p].time_state_dependent_link_from_state = [[[0 for s in range(0, g_number_of_state_vectors + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
        g_bus_list[p].time_state_dependent_link_to_state = [[[0 for s in range(0, g_number_of_state_vectors + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
        g_bus_list[p].time_state_dependent_link_cost = [[[[0 for s2 in range(0, g_number_of_state_vectors + 1)] for s1 in range(0, g_number_of_state_vectors + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
        
    global check_link_cumulative_satisfied_demand
    check_link_cumulative_satisfied_demand = [[0 for i in range(1, g_number_of_links + 1)] for l in range(1, maximum_iteration_step + 1)]
    
    global total_routing_cost
    total_routing_cost = [0 for i in range(1, maximum_iteration_step + 1)]
    
    global upper_bound
    upper_bound = [0 for i in range(1, maximum_iteration_step + 1)]
                        
    return()

    

def g_update_LR_multiplier():
    
    for l in range(1, g_number_of_links):
        change_of_LR_multiplier = check_link_cumulative_satisfied_demand[iteration_step][l]
        g_link_list[l].LR_multiplier[iteration_step + 1] = g_link_list[l].LR_multiplier[iteration_step] + step_size * change_of_LR_multiplier
                
    return()



def g_update_satisfied_demand_by_other_buses():
    
    global link_satisfied_demand_by_other_buses
    link_satisfied_demand_by_other_buses = [[0 for d in range(1, g_number_of_destinations + 1)] for l in range(1, g_number_of_links + 1)]
    
    for pp in range(1, g_number_of_buses):
        for l in range(1, g_number_of_links):   
            if (g_link_list[l].type == 1):
                start_time = g_link_list[l].start_time
                end_time = g_link_list[l].end_time
                for t in range(start_time, end_time + 1):           
                    for s in range(1, g_number_of_state_vectors):
                        if (g_bus_list[pp].time_state_dependent_link_volume[l][t][s] == 1):
                            to_state = g_bus_list[pp].time_state_dependent_link_to_state[l][t][s]
                            from_state = s
                            for d in range(1, g_number_of_destinations):
                                link_satisfied_demand_by_other_buses[l][d] = link_satisfied_demand_by_other_buses[l][d] + (g_state_vector_list[to_state].passenger_by_destination[d] - g_state_vector_list[from_state].passenger_by_destination[d])
       
    return()




def g_update_transfer_demand_generated_by_other_buses():
    
    global link_transfer_demand_generated_by_other_buses
    link_transfer_demand_generated_by_other_buses = [[[0 for d in range(1, g_number_of_destinations + 1)] for t in range(1, g_number_of_time_intervals + 1)] for l in range(1, g_number_of_links + 1)]
    
    for l in range(1, g_number_of_links):   
        if (g_link_list[l].type == 2):
            for t in range(1, g_number_of_time_intervals):
                for pp in range(1, g_number_of_buses):
                    for s in range(0, g_number_of_state_vectors):
                        if (g_bus_list[pp].time_state_dependent_link_volume[l][t][s] == 1):
                            from_state = s
                            to_state = g_bus_list[pp].time_state_dependent_link_to_state[l][t][s]
                            for d in range(1, g_number_of_destinations):
                                # need to consider
                                link_transfer_demand_generated_by_other_buses[l][t][d] = link_transfer_demand_generated_by_other_buses[l][t][d] + (g_state_vector_list[to_state].passenger_by_destination[d] - g_state_vector_list[from_state].passenger_by_destination[d])
                                    
    return()



def g_update_cumulative_satisfied_demand():
    
    global link_cumulative_satisfied_demand
    link_cumulative_satisfied_demand = [[0 for d in range(1, g_number_of_destinations + 1)] for l in range(1, g_number_of_links + 1)]
    
    for l in range(1, g_number_of_links):  
        if (g_link_list[l].type == 1):
            for pp in range(1, g_number_of_buses):
                start_time = g_link_list[l].start_time
                end_time = g_link_list[l].end_time
                for t in range(start_time, end_time + 1):         
                    for s in range(1, g_number_of_state_vectors):
                        if (g_bus_list[pp].time_state_dependent_link_volume[l][t][s] == 1):
                            to_state = g_bus_list[pp].time_state_dependent_link_to_state[l][t][s]
                            from_state = s
                            for d in range(1, g_number_of_destinations):
                                link_cumulative_satisfied_demand[l][d] = link_cumulative_satisfied_demand[l][d] + (g_state_vector_list[to_state].passenger_by_destination[d] - g_state_vector_list[from_state].passenger_by_destination[d])
                        
    return()
    

def g_check_cumulative_satisfied_demand():

    for l in range(1, g_number_of_links):
        if (g_link_list[l].type == 1):
            demand_vector_id = g_link_list[l].demand_vector_id
            for d in range(1, g_number_of_destinations):
                check_link_cumulative_satisfied_demand[iteration_step][l] += link_cumulative_satisfied_demand[l][d] - g_state_vector_list[demand_vector_id].passenger_by_destination[d]
            
    return()


def g_calcualte_total_routing_cost():
    
    global total_routing_cost
    
    for p in range(1, g_number_of_buses):
        for l in range(1, g_number_of_links):
            fixed_cost = g_link_list[l].cost
            outgoing_from_depot_flag = g_link_list[l].outgoing_from_depot_flag
            if (outgoing_from_depot_flag == 1):
                fixed_cost += usage_cost_of_reemployed_shuttles              
            for t in range(1, g_number_of_time_intervals):
                for s in range(0, g_number_of_state_vectors):
                    state_dependent_cost = g_state_vector_list[s].state_dependent_cost
                    total_routing_cost[iteration_step] += g_bus_list[p].time_state_dependent_link_volume[l][t][s] * (fixed_cost + state_dependent_cost)
    
    for p in range(1, g_number_of_buses):
        for l in range(1, g_number_of_links):
             if (g_link_list[l].type == 2):
                 for t in range(1, g_number_of_time_intervals):
                     for s in range(0, g_number_of_state_vectors):
                         if (g_bus_list[p].time_state_dependent_link_volume[l][t][s] == 1):
                             from_s = s
                             to_s = g_bus_list[p].time_state_dependent_link_to_state[l][t][s]
                             from_passegner_numnber = g_state_vector_list[from_s].total_number_of_passengers
                             to_passegner_numnber = g_state_vector_list[to_s].total_number_of_passengers
                             total_transfer_penalty = abs(from_passegner_numnber - to_passegner_numnber) * transfer_penalty
                             total_routing_cost[iteration_step] += total_transfer_penalty
             
    return()


def g_calcualte_upper_bound():
    
    global upper_bound
    
    for p in range(1, g_number_of_buses):
        for l in range(1, g_number_of_links):
            fixed_cost = g_link_list[l].cost
            outgoing_from_depot_flag = g_link_list[l].outgoing_from_depot_flag
            if (outgoing_from_depot_flag == 1):
                fixed_cost += usage_cost_of_reemployed_shuttles              
            for t in range(1, g_number_of_time_intervals):
                for s in range(0, g_number_of_state_vectors):
                    state_dependent_cost = g_state_vector_list[s].state_dependent_cost
                    upper_bound[iteration_step] += g_bus_list[p].time_state_dependent_link_volume[l][t][s] * (fixed_cost + state_dependent_cost)
    
    for p in range(1, g_number_of_buses):
        for l in range(1, g_number_of_links):
             if (g_link_list[l].type == 2):
                 for t in range(1, g_number_of_time_intervals):
                     for s in range(0, g_number_of_state_vectors):
                         if (g_bus_list[p].time_state_dependent_link_volume[l][t][s] == 1):
                             from_s = s
                             to_s = g_bus_list[p].time_state_dependent_link_to_state[l][t][s]
                             from_passegner_numnber = g_state_vector_list[from_s].total_number_of_passengers
                             to_passegner_numnber = g_state_vector_list[to_s].total_number_of_passengers
                             total_transfer_penalty = abs(from_passegner_numnber - to_passegner_numnber) * transfer_penalty
                             total_routing_cost[iteration_step] += total_transfer_penalty
    
    for l in range(1, g_number_of_links):
        if (g_link_list[l].type == 1):
            if (check_link_cumulative_satisfied_demand[iteration_step][l] < 0):
                upper_bound[iteration_step] += usage_cost_of_reemployed_shuttles + g_number_of_time_intervals
            
            
    return()


def g_write_optimal_solution():
    
    # output passenger file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'output_bus_results'    
    sheet['A1'] = 'bus'
    sheet['B1'] = 'node_sequence'
    sheet['C1'] = 'time_sequence'
    sheet['D1'] = 'state_sequence'
    sheet['E1'] = 'departure_time'
    sheet['F1'] = 'arrival_time'     
    for p in range(1, g_number_of_buses):
        row = p + 1
        departure_time = g_bus_list[p].departure_time
        arrival_time = g_bus_list[p].arrival_time
        node_sequence = ";".join(str (node) for node in g_bus_list[p].node_sequence)
        time_sequence = ";".join(str (time) for time in g_bus_list[p].time_sequence)
        state_vector_sequence = ";".join(str (state) for state in g_bus_list[p].state_vector_sequence)
        sheet.cell(row = row, column = 1, value = p)
        sheet.cell(row = row, column = 2, value = node_sequence)
        sheet.cell(row = row, column = 3, value = time_sequence)
        sheet.cell(row = row, column = 4, value = state_vector_sequence)
        sheet.cell(row = row, column = 5, value = departure_time)
        sheet.cell(row = row, column = 6, value = arrival_time)         
    workbook.save('output_bus_results.xlsx')
    
    # output check demand
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'check_demand_results' 
    sheet['A1'] = 'link_id'
    sheet['B1'] = 'demand_vector_id'
   
       

    ll = 1
    for l in range(1, g_number_of_links):        
        if (g_link_list[l].type == 1):
            row = ll + 1
            link_id = g_link_list[l].link_id
            demand = g_link_list[l].demand_vector_id
            sheet.cell(row = row, column = 1, value = link_id)
            sheet.cell(row = row, column = 2, value = demand)
            for i in range(1, maximum_iteration_step):
                sheet.cell(row = row, column = i + 2, value = check_link_cumulative_satisfied_demand[i][l])
            ll = ll + 1
            
    for i in range(1, maximum_iteration_step):
        sheet.cell(row = 1, column = i + 2, value = i)
         
    workbook.save('output_check_satisfied_demand.xlsx')
    

    return()

def g_update_member_workload():

    global g_member_workload
    global g_number_of_work_destination
    global total_member_cost

    l = 2
    g_member_workload = [0]*l
    g_number_of_work_destination = [0]*l
    total_member_cost = 0

    for p in range(1, g_number_of_buses):
        for link_element in g_bus_list[p].link_sequence:
            cost = g_link_list[link_element].cost     
            g_member_workload[p-1] = g_member_workload[p-1] + cost
            if g_link_list[link_element].type < 0:
                g_number_of_work_destination[p-1] = g_number_of_work_destination[p-1] + 1
            if g_link_list[link_element].from_node_id == g_link_list[link_element].to_node_id == 2:
                break
    for i in range (0,l):        
        total_member_cost = total_member_cost + g_member_workload[i]
    print(total_member_cost)

#    g_member_workload_std = np.std(g_member_workload)    
#    print (g_member_workload)
#    print (g_number_of_work_destination)
#    print (g_member_workload_std)
    
    return() 

if __name__=='__main__':
    
    print('Reading data......') 
    
    # define parameter
    global rou1
    global rou2
    global step_size
    global maximum_iteration_step
    global iteration_step
    global usage_cost_of_reemployed_shuttles
    global transfer_penalty
    
    maximum_iteration_step = 11
    rou1 = 1
    rou2 = 1
    usage_cost_of_reemployed_shuttles = 10
    transfer_penalty = 2
    g_read_input_data()
    g_generate_in_out_going_link()
    g_initialization()
    
    #start_Gurobi = time.time()
    #g_use_Gurobi_to_sovle_the_model()
    #end_Gurobi = time.time()
    #computing_time_Gurobi = end_Gurobi - start_Gurobi
    #print(computing_time_Gurobi)
    
    start_DP = time.time()
    for iteration_step in range(1, maximum_iteration_step):
        print(iteration_step)
        step_size = 1 / (1 + iteration_step)
        g_time_dependent_dynamic_programming_for_buses()
        g_update_cumulative_satisfied_demand()
        g_check_cumulative_satisfied_demand()
        g_update_LR_multiplier()
        g_calcualte_total_routing_cost()
        g_calcualte_upper_bound()
        g_update_member_workload()
    end_DP = time.time()
    computing_time_DP = end_DP - start_DP
    print(computing_time_DP)
    
    # output data
    g_write_optimal_solution()


    
    